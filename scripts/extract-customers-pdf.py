#!/usr/bin/env python3
"""Extract Commando Auto customer-list PDF into Nexttel POS customer import CSV/XLSX."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from pypdf import PdfReader

PDF_CANDIDATES = [
    Path(
        "/Users/sharedtechadnan/.cursor/projects/"
        "Users-sharedtechadnan-Sites-personal-KSA-disk-app-tauri-nexttel-pos/"
        "attachments/7f1067a6-9a31-4996-9f2d-02986c17af19/Customer_List.20-8-2026.pdf"
    ),
    Path.home()
    / "Documents/family documents/Abdul aziz/Customer_List.20-8-2026.pdf",
]

OUT_DIR = Path(__file__).resolve().parent.parent / "data"
CSV_PATH = OUT_DIR / "customers-cataf-import.csv"
XLSX_PATH = OUT_DIR / "customers-cataf-import.xlsx"

HEADERS = ["name", "phone", "email", "address", "notes"]

PHONE_TOKEN_RE = re.compile(
    r"^(?:\+?92|0)?3\d{8,10}$|^\+?92\d{10}$|^0\d{8,11}$|^\d{10,12}$"
)
CITY_EXPAND = {
    "DIR L TMG": "Lower Dir, Timergara",
    "DIR L": "Lower Dir",
    "DIR U": "Upper Dir",
    "DIR U DARORA": "Upper Dir, Darora",
    "DIR U WARI": "Upper Dir, Wari",
    "DIR U SHERINGAL": "Upper Dir, Sheringal",
    "DIR L KHALL": "Lower Dir, Khall",
    "DIR L KAMBAT": "Lower Dir, Kambat",
    "DIR L RABAT": "Lower Dir, Rabat",
    "DIR L SAMARBAGH": "Lower Dir, Samarbagh",
    "DIR L MUNDA": "Lower Dir, Munda",
    "DIR L MIAN BANDA": "Lower Dir, Mian Banda",
    "DIR L HAYASERI": "Lower Dir, Hayaseri",
    "DIR L KAMARTALL": "Lower Dir, Kamartall",
    "DIR L BAJAWAR": "Bajaur",
    "DIR L TMG": "Lower Dir, Timergara",
    "PESHAWAR": "Peshawar",
    "LAHOR": "Lahore",
    "PINDI": "Rawalpindi",
    "SWAT": "Swat",
    "DITR L": "Lower Dir",
    "AGHA SHIAR ALI MARKIT": "Agha Shiar Ali Market, Quetta",
    "NO": "",
}

NOISE_PREFIXES = (
    "===== page",
    "printed on",
    "page ",
    "cusotmer list",
    "showing list of customers",
    "commando auto",
    "ph:",
    "fax:",
    "mobile:",
    "web:",
    "email:",
    "w-sale delar",
    "dir lower timergarah",
)


def find_pdf() -> Path:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).expanduser()
        if not path.exists():
            raise SystemExit(f"PDF not found: {path}")
        return path
    for path in PDF_CANDIDATES:
        if path.exists():
            return path
    raise SystemExit("Customer list PDF not found. Pass the path as the first argument.")


def is_noise(line: str) -> bool:
    t = line.strip().lower()
    if not t:
        return True
    return any(t.startswith(prefix) for prefix in NOISE_PREFIXES)


def is_col_header(line: str) -> bool:
    return "customer & city" in line.lower()


def split_cols(line: str) -> list[str]:
    # Column gaps in this PDF are wide; keep 1-2 spaces inside names like "NAVID AUTOS  NKI".
    return [re.sub(r"\s+", " ", part).strip() for part in re.split(r"\s{3,}", line.strip()) if part.strip()]


def is_phone_token(token: str) -> bool:
    compact = re.sub(r"[^\d+]", "", token)
    return bool(PHONE_TOKEN_RE.match(compact))


def normalize_phone(token: str) -> str:
    digits = re.sub(r"\D", "", token)
    if digits.startswith("92") and len(digits) >= 12:
        digits = "0" + digits[2:]
    elif digits.startswith("3") and len(digits) == 10:
        digits = "0" + digits
    return digits


def normalize_name(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", name).strip()
    if cleaned.lower() == cleaned:
        cleaned = cleaned.upper()
    return cleaned


def expand_city(city: str) -> str:
    key = re.sub(r"\s+", " ", city).strip().upper()
    if key in CITY_EXPAND:
        return CITY_EXPAND[key]
    return re.sub(r"\s+", " ", city).strip()


def clean_place(text: str) -> str:
    value = re.sub(r"\s+", " ", text).strip()
    replacements = {
        "TIMERGARAH": "Timergara",
        "TIMERGARA": "Timergara",
        "TIMERGRAH": "Timergara",
        "TIIMERGARA": "Timergara",
        "LAHOR": "Lahore",
        "BADAMIBAGH": "Badami Bagh",
        "BADAMI BAGH": "Badami Bagh",
        "SHERINAGAL": "Sheringal",
        "CHUKYATN": "Chukiatan",
        "CHUKYATIN": "Chukiatan",
        "CHUKYATAN": "Chukiatan",
        "TIMRAGARAH": "Timergara",
        "AJ SHUBA": "Ajj Shoba",
        "AJJ SHOBA": "Ajj Shoba",
        "ITIPAQ SHUBA": "Ittifaq Shoba",
        "KHALL SHOBA": "Khall Shoba",
        "KIMP ADA": "Kimp Ada",
        "SADAR BAZAR": "Saddar Bazaar",
        "SHUBA BAZAR": "Shoba Bazaar",
        "SHOBA BAZAR": "Shoba Bazaar",
        "AUTO CENTER": "Auto Center",
        "AUTO DICORATION": "Auto Decoration",
        "MIAN ADA": "Mian Ada",
        "MIAN BANDA": "Mian Banda",
        "BAJAWAR": "Bajaur",
    }
    result = value
    for src, dest in replacements.items():
        result = re.sub(re.escape(src), dest, result, flags=re.I)
    words = []
    for word in re.split(r"\s+", result.strip()):
        if word.isupper() and word not in {"NKI", "NTC", "O/S"} and len(word) > 1:
            words.append(word.title())
        else:
            words.append(word)
    return " ".join(words).strip()


def unique_join(*chunks: str) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        if not chunk:
            continue
        for piece in re.split(r",\s*", chunk):
            piece = piece.strip(" ,")
            if not piece:
                continue
            key = piece.lower()
            if key in seen:
                continue
            joined = " ".join(parts).lower()
            if joined and re.search(rf"\b{re.escape(key)}\b", joined):
                continue
            seen.add(key)
            parts.append(piece)
    return ", ".join(parts)


def build_address(local: str, city: str) -> str:
    local_clean = clean_place(local) if local and local.upper() != "NO" else ""
    city_pretty = expand_city(city)
    if local_clean and re.search(r"\bDir\b", city_pretty, re.I):
        stripped = re.sub(r"\s*Dir$", "", local_clean, flags=re.I).strip()
        if stripped and stripped.lower() not in {"upper", "lower"}:
            local_clean = stripped
    return unique_join(local_clean, city_pretty)


def parse_line_parts(line: str | None) -> tuple[str, list[str], str]:
    if not line:
        return "", [], ""
    parts = split_cols(line)
    if not parts:
        return "", [], ""
    lead = parts[0]
    phones: list[str] = []
    other: list[str] = []
    for token in parts[1:]:
        if is_phone_token(token):
            phones.append(normalize_phone(token))
        else:
            other.append(token)
    return lead, phones, " ".join(other).strip()


def is_walk_in(name: str) -> bool:
    return bool(re.match(r"^walk[\s-]*in", name.strip(), re.I))


def extract_layout_lines(pdf_path: Path) -> list[str]:
    reader = PdfReader(str(pdf_path))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text(extraction_mode="layout") or ""
        for raw in text.splitlines():
            line = raw.rstrip()
            if is_noise(line):
                continue
            lines.append(line)
    return lines


def next_index(lines: list[str], start: int) -> int:
    i = start
    while i < len(lines) and not lines[i].strip():
        i += 1
    return i


def is_section_heading(lines: list[str], index: int) -> bool:
    nxt = next_index(lines, index + 1)
    return nxt < len(lines) and is_col_header(lines[nxt])


def parse_customers(lines: list[str]) -> list[dict]:
    customers: list[dict] = []
    i = 0
    n = len(lines)
    while i < n:
        i = next_index(lines, i)
        if i >= n:
            break
        line = lines[i]
        if is_col_header(line) or is_section_heading(lines, i):
            i += 1
            continue

        name_line = line
        city_line = None
        j = next_index(lines, i + 1)
        if j < n and not is_col_header(lines[j]) and not is_section_heading(lines, j):
            city_line = lines[j]
            i = j + 1
        else:
            i += 1

        name, name_phones, local_address = parse_line_parts(name_line)
        city, city_phones, city_extra = parse_line_parts(city_line)
        name = normalize_name(name)
        if not name or is_walk_in(name):
            continue
        if city.upper() == "NO":
            city = ""

        phones: list[str] = []
        for phone in name_phones + city_phones:
            if phone and phone not in phones:
                phones.append(phone)

        extra_place = local_address or city_extra
        address = build_address(extra_place, city)
        notes_parts = []
        if city:
            notes_parts.append(f"Area: {re.sub(r'\s+', ' ', city).strip()}")
        notes = " | ".join(notes_parts)

        customers.append(
            {
                "name": name,
                "phone": " / ".join(phones),
                "email": "",
                "address": address,
                "notes": notes,
            }
        )
    return customers


def write_csv(rows: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row[key] for key in HEADERS])
    widths = {"A": 36, "B": 36, "C": 22, "D": 48, "E": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    instructions = wb.create_sheet("Instructions")
    instructions.append(["Nexttel POS customer import"])
    instructions.append(["Required column: name"])
    instructions.append(["Optional columns: phone, email, address, notes"])
    instructions.append(["Import via Customers → Import"])
    instructions.append(["Walk-in / cash-sale customers are omitted; use Walk-in on POS instead."])
    wb.save(XLSX_PATH)


def main() -> None:
    pdf_path = find_pdf()
    lines = extract_layout_lines(pdf_path)
    rows = parse_customers(lines)
    write_csv(rows)
    write_xlsx(rows)
    with_phone = sum(1 for row in rows if row["phone"])
    print(f"PDF: {pdf_path}")
    print(f"Customers: {len(rows)}")
    print(f"With phone: {with_phone}")
    print(f"Without phone: {len(rows) - with_phone}")
    print(f"Wrote {CSV_PATH}")
    print(f"Wrote {XLSX_PATH}")
    print("\nFirst 8:")
    for row in rows[:8]:
        print(f"  {row['name']} | {row['phone']} | {row['address']}")
    print("\nLast 5:")
    for row in rows[-5:]:
        print(f"  {row['name']} | {row['phone']} | {row['address']}")


if __name__ == "__main__":
    main()
