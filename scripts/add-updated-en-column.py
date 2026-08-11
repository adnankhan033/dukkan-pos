#!/usr/bin/env python3
"""Add updated_en column by translating name_ar to proper English."""
import csv
import importlib.util
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent

spec = importlib.util.spec_from_file_location("fix_export", SCRIPT_DIR / "fix-products-export.py")
fx = importlib.util.module_from_spec(spec)
spec.loader.exec_module(fx)

has_arabic = fx.has_arabic
clean_arabic = fx.clean_arabic
clean_english_from_arabic = fx.clean_english_from_arabic
title_english = fx.title_english
is_bad_english = fx.is_bad_english


def translate_name_ar(name_ar, fallback_name=""):
    text = str(name_ar or "").strip()
    fallback = str(fallback_name or "").strip()

    if has_arabic(text):
        return clean_english_from_arabic(clean_arabic(text))

    if text and not is_bad_english(text):
        return title_english(text)

    if fallback and not has_arabic(fallback):
        return title_english(fallback)

    return title_english(text or fallback)


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    data = []
    for values in rows:
        row = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
        if any(v is not None and str(v).strip() for v in row.values()):
            data.append(row)
    wb.close()
    return headers, data


def output_headers(original_headers):
    headers = [h for h in original_headers if h != "updated_en"]
    if "name_ar" in headers:
        idx = headers.index("name_ar") + 1
        headers.insert(idx, "updated_en")
    else:
        headers.append("updated_en")
    return headers


def write_csv(path, headers, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def main():
    src = Path("/Users/sharedtechadnan/Downloads/products-export-20260811-1654.xlsx")
    out_dir = ROOT / "data"
    out_dir.mkdir(exist_ok=True)
    csv_out = out_dir / "products-with-updated-en.csv"
    xlsx_out = out_dir / "products-with-updated-en.xlsx"
    downloads_xlsx = Path("/Users/sharedtechadnan/Downloads/products-with-updated-en.xlsx")
    downloads_csv = Path("/Users/sharedtechadnan/Downloads/products-with-updated-en.csv")

    if not src.exists():
        print(f"Source not found: {src}", file=sys.stderr)
        sys.exit(1)

    original_headers, raw = read_xlsx(src)
    headers = output_headers(original_headers)

    out_rows = []
    from_ar = 0
    from_en = 0
    for row in raw:
        name_ar = row.get("name_ar", "")
        updated_en = translate_name_ar(name_ar, row.get("name", ""))
        if has_arabic(str(name_ar or "")):
            from_ar += 1
        else:
            from_en += 1
        out = dict(row)
        out["updated_en"] = updated_en
        out_rows.append(out)

    write_csv(csv_out, headers, out_rows)
    write_xlsx(xlsx_out, headers, out_rows)

    try:
        shutil.copy2(xlsx_out, downloads_xlsx)
        shutil.copy2(csv_out, downloads_csv)
    except OSError as exc:
        print(f"Note: could not copy to Downloads ({exc})")

    print(f"Processed {len(out_rows)} products")
    print(f"Translated from Arabic name_ar: {from_ar}")
    print(f"Kept English (no Arabic in name_ar): {from_en}")
    print(f"CSV:  {csv_out}")
    print(f"XLSX: {xlsx_out}")
    if downloads_xlsx.exists():
        print(f"Copy: {downloads_xlsx}")

    print("\nSample rows:")
    for row in out_rows[:5]:
        print(f"  name_ar: {str(row.get('name_ar', ''))[:55]}")
        print(f"  updated_en: {row['updated_en']}")
        print(f"  name (old): {str(row.get('name', ''))[:55]}")
        print()


if __name__ == "__main__":
    main()
