#!/usr/bin/env python3
"""Clean product export: fix English/Arabic names and categories for POS import."""
import csv
import importlib.util
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent

spec = importlib.util.spec_from_file_location("convert_zoho", SCRIPT_DIR / "convert-zoho-items.py")
cz = importlib.util.module_from_spec(spec)
spec.loader.exec_module(cz)

HEADERS = cz.HEADERS
has_arabic = cz.has_arabic
normalize_ar = cz.normalize_ar
infer_category = cz.infer_category
detect_unit = cz.detect_unit
is_bad_english = cz.is_bad_english

EXTRA_BRANDS = [
    ("المرعي", "Almarai"), ("المراعي", "Almarai"), ("ام&امز", "M&M"), ("ام&ام", "M&M"),
    ("ام & ام", "M&M"), ("أومبا", "Ompa"), ("اورالبى", "Oral-B"), ("اورالبي", "Oral-B"),
    ("فاين", "Fine"), ("سانيتا", "Sanita"), ("سانتيا", "Sanita"), ("بامبي", "Pampers"),
    ("بيبي فاين", "Fine Baby"), ("بيبيلاك", "Bebelac"), ("حدائق", "Hadaeq"), ("شتورة", "Shatoura"),
    ("السيف", "Al Saif"), ("تايجر", "Tiger"), ("الكاس", "Al Kas"), ("اولكر", "Ulker"),
    ("7 أيام", "7Days"), ("راجا", "Raja"), ("سنكرس", "Snickers"), ("امريكانا", "Americana"),
    ("سبعة نجوم", "Seven Stars"), ("داك", "Dac"), ("روكسونا", "Rexona"), ("مهران", "Mehran"),
    ("بيبلو", "Bibo"), ("ديو", "Dove"), ("جوليت", "Gillette"), ("النجم", "Al Najm"),
    ("لونا", "Luna"), ("بزاليا", "Luna"), ("بيوقلز", "Bugles"), ("ابوسهم", "Abu Shum"),
    ("ابو سهم", "Abu Shum"), ("Radiant", "Radiant"), ("radiant", "Radiant"),
]

EXTRA_TERMS = [
    ("مشروب طاقة", "Energy Drink"), ("مشروب", "Drink"), ("عصير", "Juice"), ("حليب", "Milk"),
    ("لبن", "Laban"), ("جبن", "Cheese"), ("جبنة", "Cheese"), ("زبادي", "Yogurt"), ("زبدة", "Butter"),
    ("سمن", "Ghee"), ("بيض", "Eggs"), ("مياه", "Water"), ("ماء", "Water"), ("شاي", "Tea"),
    ("قهوة", "Coffee"), ("كولا", "Cola"), ("بسكويت", "Biscuits"), ("شوكolat", "Chocolate"),
    ("شokolat", "Chocolate"), ("شوكلاتة", "Chocolate"), ("حلاوة", "Halawa"), ("حلاوه", "Halawa"),
    ("علكة", "Chewing Gum"), ("لبان", "Chewing Gum"), ("شips", "Chips"), ("شيبس", "Chips"),
    ("تونة", "Tuna"), ("تونه", "Tuna"), ("معجون", "Paste"), ("طماطم", "Tomato"),
    ("كاتchup", "Ketchup"), ("كاتشب", "Ketchup"), ("مايونيز", "Mayonnaise"), ("صلصة", "Sauce"),
    ("زيت", "Oil"), ("أرز", "Rice"), ("ارز", "Rice"), ("دقيق", "Flour"), ("سكر", "Sugar"),
    ("ملح", "Salt"), ("بهارات", "Spices"), ("فلفل", "Pepper"), ("كمون", "Cumin"),
    ("خبz", "Bread"), ("خبز", "Bread"), ("كيك", "Cake"), ("دجاج", "Chicken"), ("لحم", "Meat"),
    ("مفروم", "Mince"), ("صابون", "Soap"), ("شامبو", "Shampoo"), ("معجون اسنان", "Toothpaste"),
    ("معجون أسنان", "Toothpaste"), ("فرشaة", "Toothbrush"), ("فرشة", "Toothbrush"),
    ("فرشاة أسنان", "Toothbrush"), ("حفاض", "Diapers"), ("حفائظ", "Diapers"), ("حفاضات", "Diapers"),
    ("فوط", "Pads"), ("صحية", "Sanitary"), ("منadil", "Tissue"), ("مناديل", "Tissue"),
    ("ورق", "Paper"), ("منظف", "Cleaner"), ("مسحوق", "Detergent"), ("غسيل", "Laundry"),
    ("بطاريات", "Batteries"), ("بطارية", "Battery"), ("لمبة", "Bulb"), ("ولاعة", "Lighter"),
    ("فحم", "Charcoal"), ("مخلل", "Pickles"), ("كبيس", "Pickles"), ("طرشي", "Pickles"),
    ("فول", "Fava Beans"), ("فول سوداني", "Peanuts"), ("حمص", "Chickpeas"), ("تمر", "Dates"),
    ("لوz", "Almonds"), ("لوز", "Almonds"), ("فستق", "Pistachio"), ("شmer", "Fennel"),
    ("شمر", "Fennel"), ("بخور", "Incense"), ("قفازات", "Gloves"), ("مكينة", "Razor"),
    ("حلاقة", "Shaving"), ("صبغة", "Hair Dye"), ("شعر", "Hair"), ("معقم", "Antiseptic"),
    ("مطهر", "Disinfectant"), ("بطاطس", "Potatoes"), ("بصل", "Onion"), ("خيار", "Cucumber"),
    ("موز", "Banana"), ("تفاح", "Apple"), ("برتقال", "Orange"), ("بندق", "Hazelnut"),
    ("فراولة", "Strawberry"), ("مانgo", "Mango"), ("مانجو", "Mango"), ("مهلبية", "Pudding"),
    ("قشطة", "Cream"), ("كريمة", "Cream"), ("مفرش", "Table Cover"), ("لاصق", "Tape"),
    ("سكين", "Knife"), ("مكعبات", "Cubes"), ("مرق", "Stock"), ("مكرونة", "Pasta"),
    ("معmoul", "Maamoul"), ("معمول", "Maamoul"), ("فاصوليا", "Beans"), ("عدس", "Lentils"),
    ("ذرة", "Corn"), ("بازilla", "Peas"), ("بسلة", "Peas"), ("زيتون", "Olives"), ("طحينية", "Tahini"),
    ("سائل", "Liquid"), ("محول", "Adapter"), ("كهربائي", "Electric"), ("ثلاجة", "Cooler"),
    ("سخان", "Heater"), ("توصيلة", "Extension Cord"), ("سلك", "Scrubber"), ("ستانلس", "Stainless"),
    ("قشارة", "Peeler"), ("قدر", "Pot"), ("قدور", "Pots"), ("صواني", "Trays"), ("زجاج", "Glass"),
    ("مزيل عرق", "Deodorant"), ("مزيل", "Remover"), ("بلسم", "Conditioner"), ("أصابع", "Fingers"),
    ("سمك", "Fish"), ("خميرة", "Yeast"), ("خميرية", "Yeast"), ("فورية", "Instant"),
    ("سائل غسيل", "Laundry Liquid"), ("عبايات", "Abaya"), ("ملون", "Colored"), ("مشكل", "Mixed"),
    ("حار", "Hot"), ("اخضر", "Green"), ("أخضر", "Green"), ("برائحة", "Scented"),
    ("لافندر", "Lavender"), ("مغطى", "Coated"), ("وسط", "Medium"), ("كبير", "Large"),
    ("صغير", "Small"), ("كيس", "Pack"), ("علبة", "Can"), ("حبة", "Pcs"), ("العاب", "Toys"),
    ("اطفal", "Kids"), ("أطفal", "Kids"), ("اطفال", "Kids"), ("مكسرات", "Nuts"), ("بطعم", "Flavored"),
    ("المشوي", "BBQ"), ("شراب", "Syrup"), ("فرنسي", "French"), ("خوخ", "Peach"),
    ("الجوافة", "Guava"), ("بسمتي", "Basmati"), ("كرnl", "Kernel"), ("كرنل", "Kernel"),
    ("معلب", "Canned"), ("معلبة", "Canned"), ("ناget", "Nuggets"), ("برger", "Burger"),
    ("برجر", "Burger"), ("آيس كريم", "Ice Cream"), ("آيس", "Ice"), ("كورn", "Corn"),
    ("كورن", "Corn"), ("فليكس", "Flakes"), ("موس", "Peeler"), ("استيل", "Steel"),
    ("بيد", "Handle"), ("خشبي", "Wood"), ("ريال", "Riyal"), ("نكتار", "Nectar"),
]

AR_TYPOS = {
    "المرعي": "المراعي", "حدئق": "حدائق", "جفائظ": "حفاضات", "حفائض": "حفاضات",
    "تونه": "تونة", "حلاوه": "حلاوة", "مياة": "مياه", "برينجيلز": "برينجلز",
    "كوديرت": "كودريت", "ام&امز": "M&M", "ام&ام": "M&M", "اورالبى": "أورال-بي",
    "اورالبي": "أورال-بي",
}

EXTRA_CATEGORY_RULES = [
    ("Kitchen & Home", ["قدr", "قدور", "قدر", "pot", "pots", "صواني", "tray", "قشارة", "peeler",
                        "سلك", "scrubber", "ستانلس", "stainless", "مفرش", "table cover", "طبق",
                        "سخان", "heater", "ثلاجة", "cooler", "موصل", "adapter", "socket", "outlet",
                        "توصيلة", "extension", "ولاعة", "lighter", "ثقاب", "matches", "أكواب",
                        "cups", "plate", "أطباق", "foil", "aluminum", "nylon", "cling"]),
    ("Electronics", ["2800w", "watt", "كهرب", "electric", "bulb", "لمبة", "battery", "بطار"]),
    ("Toys & Games", ["العاب", "toys", "لعبة", "game"]),
    ("Health & Pharmacy", ["vitamin", "دواء", "medicine", "pharmacy", "مسكن", "pain"]),
]

JUNK_PREFIX_RE = re.compile(r'^[\s!\"#%&*\'()+./\\\-]+')
SIZE_EXTRACT_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(?:\*(\d+(?:\.\d+)?))?\s*"
    r"(مل|ml|ML|م\s*ل|لتر|ل\b|L\b|lt|LT|جم|ج\b|g\b|G\b|gram|grams|gr\b|غم|غرام|جرام|ك\b|كجم|كيلo|kg|KG|م\b|متر|m\b|ح\b|حبة|حبه|pcs|piece|pieces)?",
    re.IGNORECASE,
)


def infer_category_extended(name_en, name_ar):
    combined = f"{name_en} {name_ar}".lower()
    for category, keywords in EXTRA_CATEGORY_RULES:
        for kw in keywords:
            if kw.lower() in combined:
                return category
    return infer_category(name_en, name_ar)


def fix_ar_typos(text):
    t = text
    for old, new in AR_TYPOS.items():
        t = t.replace(old, new)
    return t


def normalize_arabic_units(text):
    if not text:
        return ""
    t = fix_ar_typos(normalize_ar(str(text).strip()))
    t = JUNK_PREFIX_RE.sub("", t)
    t = re.sub(r"(\d)(مل|جم|غم|لتر|كجم|م)(?!\s)", r"\1 \2", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def extract_sizes(text):
    sizes = []
    unit_map = {
        "مل": "ml", "ml": "ml", "م ل": "ml", "لتر": "L", "ل": "L", "lt": "L",
        "جم": "g", "ج": "g", "g": "g", "gram": "g", "grams": "g", "gr": "g",
        "غم": "g", "غرام": "g", "جرام": "g", "كجم": "kg", "ك": "kg", "كيلo": "kg", "kg": "kg",
        "م": "m", "متر": "m", "ح": "pcs", "حبة": "pcs", "حبه": "pcs",
    }
    for m in SIZE_EXTRACT_RE.finditer(text or ""):
        n1, n2, unit = m.group(1), m.group(2), (m.group(3) or "").lower().replace(" ", "")
        u = unit_map.get(unit, "")
        if n2:
            sizes.append(f"{n1}x{n2}{u}")
        elif u:
            sizes.append(f"{n1}{u}")
        elif n1 and not sizes:
            sizes.append(n1)
    return sizes


def strip_junk_english(text):
    t = re.sub(r"\s+", " ", str(text or "").strip())
    t = JUNK_PREFIX_RE.sub("", t)
    t = re.sub(r"^[\(\)\[\]\*]+\s*", "", t)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t


def all_brands():
    seen = set()
    out = []
    for ar, en in sorted(EXTRA_BRANDS + cz.BRANDS, key=lambda x: -len(x[0])):
        key = (ar.lower(), en.lower())
        if key not in seen:
            seen.add(key)
            out.append((ar, en))
    return out


def all_terms():
    seen = set()
    out = []
    for ar, en in sorted(EXTRA_TERMS + cz.TERM_EN, key=lambda x: -len(x[0])):
        if ar not in seen:
            seen.add(ar)
            out.append((ar, en))
    return out


def clean_arabic(text):
    return normalize_arabic_units(text)


def clean_english_from_arabic(ar_text):
    ar_clean = normalize_arabic_units(ar_text)
    sizes = extract_sizes(ar_clean)
    work = ar_clean
    for ar, en in all_brands():
        work = re.sub(re.escape(ar), en, work, flags=re.IGNORECASE)
    for ar, en in all_terms():
        work = work.replace(ar, en)
    work = re.sub(r"[\u0600-\u06FF]+", " ", work)
    work = re.sub(r"[^\w\s+\-*./&]", " ", work)
    work = re.sub(r"\s+", " ", work).strip()
    if is_bad_english(work):
        work = cz.phonetic(ar_clean)
        for ar, en in all_brands():
            work = re.sub(re.escape(ar), en, work, flags=re.IGNORECASE)
        for ar, en in all_terms():
            work = work.replace(ar, en)
        work = re.sub(r"[\u0600-\u06FF]+", " ", work)
        work = re.sub(r"\s+", " ", work).strip()
    parts = [p for p in work.split() if p and p not in {"*", "+", "-", "&", "Pcs", "pcs"}]
    if sizes:
        size_str = " ".join(sizes[-2:])
        parts = [p for p in parts if not re.fullmatch(r"\d+(\.\d+)?", p)]
        if size_str not in " ".join(parts):
            parts.append(size_str)
    result = strip_junk_english(" ".join(parts))
    words = []
    for w in result.split():
        if re.fullmatch(r"\d+(\.\d+)?(ml|l|g|kg|m|pcs|x\d+)?", w, re.I):
            words.append(w)
        elif w.isupper() and len(w) <= 5:
            words.append(w)
        elif w.lower() in {"and", "of", "with", "free"}:
            words.append(w.lower())
        elif "-" in w:
            words.append(w)
        else:
            words.append(w[:1].upper() + w[1:] if len(w) > 1 else w.upper())
    return strip_junk_english(" ".join(words))


def title_english(text):
    text = re.sub(r"\s+", " ", str(text or "").strip())
    if not text:
        return ""
    words = []
    for w in text.split():
        if re.fullmatch(r"\d+(\.\d+)?(ml|l|g|kg|gm|gr|pcs)?", w, re.I):
            words.append(w)
        elif w.isupper() and len(w) <= 5:
            words.append(w)
        else:
            words.append(w[:1].upper() + w[1:] if len(w) > 1 else w.upper())
    return strip_junk_english(" ".join(words))


def fix_names(name, name_ar):
    name = str(name or "").strip()
    name_ar = str(name_ar or "").strip()

    if has_arabic(name) and not has_arabic(name_ar) and re.search(r"[a-zA-Z]", name_ar):
        name, name_ar = name_ar, name

    if has_arabic(name_ar):
        ar = clean_arabic(name_ar)
        en = clean_english_from_arabic(ar)
        if is_bad_english(en) and name and not has_arabic(name):
            en = title_english(name)
        return en, ar

    if has_arabic(name):
        ar = clean_arabic(name)
        en = clean_english_from_arabic(ar)
        return en, ar

    en = title_english(name or name_ar)
    ar = en
    if re.search(r"riyal|sar", en, re.I):
        ar = en.replace("Riyal", "ريال").replace("riyal", "ريال")
    return en, ar


def fix_row(row):
    name, name_ar = fix_names(row.get("name"), row.get("name_ar"))
    category = infer_category_extended(name, name_ar)
    unit = detect_unit(f"{name} {name_ar}") or row.get("unit") or "pcs"

    return {
        "name": name,
        "name_ar": name_ar,
        "sku": row.get("sku") or "",
        "barcode": row.get("barcode") or "",
        "category": category,
        "unit": unit,
        "supplier": row.get("supplier") or "",
        "cost_price": row.get("cost_price") if row.get("cost_price") is not None else "0.00",
        "selling_price": row.get("selling_price") if row.get("selling_price") is not None else "0.00",
        "quantity": row.get("quantity") if row.get("quantity") is not None else "0",
        "min_stock": row.get("min_stock") if row.get("min_stock") is not None else "5",
        "published": row.get("published") or "yes",
    }


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    data = []
    for values in rows:
        row = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
        if any(v is not None and str(v).strip() for v in row.values()):
            data.append(row)
    wb.close()
    return data


def write_csv(path, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row[h] for h in HEADERS])
    wb.save(path)


def main():
    src = Path("/Users/sharedtechadnan/Downloads/products-export-20260811-1654.xlsx")
    out_dir = ROOT / "data"
    out_dir.mkdir(exist_ok=True)
    csv_out = out_dir / "products-cleaned-import.csv"
    xlsx_out = out_dir / "products-cleaned-import.xlsx"
    downloads_out = Path("/Users/sharedtechadnan/Downloads/products-cleaned-import.xlsx")

    if not src.exists():
        print(f"Source not found: {src}", file=sys.stderr)
        sys.exit(1)

    raw = read_xlsx(src)
    fixed = [fix_row(r) for r in raw]

    write_csv(csv_out, fixed)
    write_xlsx(xlsx_out, fixed)
    try:
        shutil.copy2(xlsx_out, downloads_out)
        shutil.copy2(csv_out, downloads_out.with_suffix(".csv"))
    except OSError as exc:
        print(f"Note: could not copy to Downloads ({exc})")
        downloads_out = None

    cats = {}
    changed = 0
    for before, after in zip(raw, fixed):
        if (str(before.get("name", "")).strip() != after["name"]
                or str(before.get("name_ar", "")).strip() != after["name_ar"]
                or str(before.get("category", "")).strip() != after["category"]):
            changed += 1
        cats[after["category"]] = cats.get(after["category"], 0) + 1

    print(f"Processed {len(fixed)} products")
    print(f"Updated {changed} rows")
    print(f"CSV:  {csv_out}")
    print(f"XLSX: {xlsx_out}")
    if downloads_out:
        print(f"Copy: {downloads_out}")
    print("Categories:", dict(sorted(cats.items(), key=lambda x: -x[1])))


if __name__ == "__main__":
    main()
